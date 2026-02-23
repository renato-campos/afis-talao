import tkinter as tk
import csv
import logging
import os
import sys
import webbrowser
from datetime import date, datetime, time
from pathlib import Path
from tempfile import gettempdir
from tkinter import filedialog, messagebox, ttk

try:
    import customtkinter as ctk
except ImportError:
    ctk = None

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

from .constants import (
    EDITABLE_FIELDS,
    FIELD_LABELS,
    STATUS_CANCELADO,
    STATUS_FINALIZADO,
    STATUS_MONITORADO,
    STATUS_OPCOES,
)
from .config import get_env
from .interfaces import TalaoRepository
from .repository import ConcurrencyError, DuplicateTalaoError
from .services import AlertaService, TalaoService

logger = logging.getLogger(__name__)

UI_THEME = {
    "bg": "#585858",
    "header": "#1941B7",
    "surface": "#F4F2DD",
    "surface_alt": "#EEF3FA",
    "surface_hover": "#DFE7F2",
    "text": "#0F172A",
    "muted": "#334155",
    "primary": "#0B63CE",
    "primary_hover": "#0A54AE",
    "success": "#1F8A4D",
    "success_hover": "#18703E",
    "warning": "#F57C00",
    "warning_hover": "#E06F00",
    "neutral": "#E2E8F0",
    "neutral_hover": "#CBD5E1",
    "border": "#D5DEEA",
    "placeholder": "#94A3B8",
    "white": "#FFFFFF",
    "status_monitorado_bg": "#F8E171",
    "status_monitorado_fg": "#6B5600",
    "status_finalizado_bg": "#EAF7EE",
    "status_finalizado_fg": "#1E5E36",
    "status_cancelado_bg": "#FDECEC",
    "status_cancelado_fg": "#7F1D1D",
    "danger": "#7F1D1D",
    "danger_hover": "#6B1111",
}
BUTTON_FONT_BOLD = ("Segoe UI", 11, "bold")
WATERMARK_MAX_WIDTH = 2560
WATERMARK_MAX_HEIGHT = 1440
HELP_ICON_SIZE_PX = 24
WHATSAPP_ICON_SIZE_PX = 56

ALERT_INTERVAL_OPTIONS = [
    ("1 min", 1),
    ("5 min", 5),
    ("15 min", 15),
    ("30 min (padrão)", 30),
    ("60 min", 60),
]
DEFAULT_ALERT_INTERVAL_MIN = 30
DEFAULT_ALERT_INTERVAL_LABEL = next(
    (label for label, minutes in ALERT_INTERVAL_OPTIONS if minutes == DEFAULT_ALERT_INTERVAL_MIN),
    ALERT_INTERVAL_OPTIONS[0][0],
)


def format_talao(ano, numero):
    """Formata o talao no padrao NNNN/AAAA."""
    try:
        return f"{int(numero):04d}/{int(ano)}"
    except (TypeError, ValueError):
        return f"{numero or '----'}/{ano or '----'}"


def _normalize_user_text(value, field_name):
    """Padroniza texto digitado pelo usuario antes da validacao."""
    if value is None:
        return ""
    text = str(value).strip()
    if field_name in ("observacao", "status"):
        return text
    return text.upper()


def _apply_toplevel_theme(window, bg_key="bg"):
    """Aplica paleta visual em janelas toplevel Tk/CTk."""
    if ctk is not None and isinstance(window, ctk.CTkToplevel):
        window.configure(fg_color=UI_THEME[bg_key])
    else:
        window.configure(bg=UI_THEME[bg_key])


def _build_button(parent, text, command, variant="neutral", use_ctk=False, width=None):
    """Cria botao padronizado para Tk e CustomTkinter."""
    cfg = {
        "primary": (UI_THEME["primary"], UI_THEME["primary_hover"], UI_THEME["white"]),
        "success": (UI_THEME["success"], UI_THEME["success_hover"], UI_THEME["white"]),
        "warning": (UI_THEME["warning"], UI_THEME["warning_hover"], UI_THEME["white"]),
        "danger": (UI_THEME["danger"], UI_THEME["danger_hover"], UI_THEME["white"]),
        "neutral": (UI_THEME["neutral"], UI_THEME["neutral_hover"], UI_THEME["text"]),
    }
    fg, hover, text_color = cfg.get(variant, cfg["neutral"])
    if use_ctk and ctk is not None:
        button_kwargs = {
            "text": text,
            "command": command,
            "fg_color": fg,
            "hover_color": hover,
            "text_color": text_color,
            "font": BUTTON_FONT_BOLD,
            "corner_radius": 8,
            "height": 34,
        }
        if width is not None:
            button_kwargs["width"] = width
        return ctk.CTkButton(parent, **button_kwargs)
    return tk.Button(
        parent,
        text=text,
        command=command,
        bg=fg,
        fg=text_color,
        font=BUTTON_FONT_BOLD,
        relief="flat",
        activebackground=hover,
        activeforeground=text_color,
    )


def _center_toplevel_on_parent(window, parent):
    """Centraliza uma janela filha em relacao a janela pai."""
    window.update_idletasks()
    parent.update_idletasks()

    width = window.winfo_width()
    height = window.winfo_height()
    if width <= 1 or height <= 1:
        width = window.winfo_reqwidth()
        height = window.winfo_reqheight()

    parent_x = parent.winfo_rootx()
    parent_y = parent.winfo_rooty()
    parent_w = parent.winfo_width()
    parent_h = parent.winfo_height()

    x = parent_x + (parent_w - width) // 2
    y = parent_y + (parent_h - height) // 2
    window.geometry(f"+{max(0, x)}+{max(0, y)}")


class TalaoEditor(tk.Toplevel):
    """Janela modal de edicao de talao."""

    def __init__(
        self,
        parent,
        repo: TalaoRepository,
        talao_service: TalaoService,
        alerta_service: AlertaService,
        talao_id,
        intervalo_min,
        on_saved,
    ):
        super().__init__(parent)
        self.repo = repo
        self.talao_service = talao_service
        self.alerta_service = alerta_service
        self.talao_id = talao_id
        self.on_saved = on_saved
        self.title("Editar Talão")
        self.geometry("450x610")
        self.minsize(450, 610)
        self.resizable(True, True)
        self.use_ctk = ctk is not None

        self.widgets = {}
        self.intervalo_var = tk.StringVar(value=str(intervalo_min))
        self.data_bo_placeholder_active = False
        self.form_parent = self

        _apply_toplevel_theme(self, bg_key="surface")

        row = 0
        record = self.repo.get_talao(talao_id)
        if not record:
            messagebox.showerror("Erro", "Talão não encontrado.")
            self.destroy()
            return
        self.record = record

        if self.use_ctk:
            self.form_parent = ctk.CTkFrame(
                self,
                fg_color=UI_THEME["surface"],
                corner_radius=12,
                border_width=1,
                border_color=UI_THEME["border"],
            )
            self.form_parent.pack(fill="both", expand=True, padx=14, pady=14)
            ctk.CTkLabel(
                self.form_parent,
                text=f"Editar Talão {format_talao(record.get('ano'), record.get('talao'))}",
                font=("Segoe UI", 18, "bold"),
                text_color=UI_THEME["text"],
            ).grid(row=row, column=0, columnspan=2, sticky="w", padx=14, pady=(14, 12))
        else:
            tk.Label(
                self,
                text=f"Talão {format_talao(record.get('ano'), record.get('talao'))}",
                font=("Segoe UI", 12, "bold"),
                bg=UI_THEME["surface"],
                fg=UI_THEME["text"],
            ).grid(
                row=row, column=0, columnspan=2, sticky="w", padx=12, pady=(10, 12)
            )
        row += 1

        for key in EDITABLE_FIELDS:
            label_text = FIELD_LABELS[key]
            if key == "data_bo":
                label_text = "Data BO"
            if self.use_ctk:
                ctk.CTkLabel(
                    self.form_parent,
                    text=label_text,
                    font=("Segoe UI", 12),
                    text_color=UI_THEME["muted"],
                ).grid(row=row, column=0, sticky="w", padx=14, pady=4)
            else:
                tk.Label(
                    self,
                    text=label_text,
                    bg=UI_THEME["surface"],
                    fg=UI_THEME["muted"],
                    font=("Segoe UI", 10),
                ).grid(row=row, column=0, sticky="w", padx=12, pady=4)

            if key == "status":
                if self.use_ctk:
                    widget = ctk.CTkComboBox(
                        self.form_parent,
                        values=list(STATUS_OPCOES),
                        variable=tk.StringVar(value=str(record.get(key) or STATUS_MONITORADO)),
                        state="readonly",
                        width=430,
                    )
                else:
                    widget = ttk.Combobox(self, values=STATUS_OPCOES, state="readonly")
                widget.set(str(record.get(key) or STATUS_MONITORADO))
            elif key in ("observacao", "vitimas"):
                if self.use_ctk:
                    widget = ctk.CTkTextbox(
                        self.form_parent,
                        width=430,
                        height=96 if key == "observacao" else 72,
                        fg_color=UI_THEME["surface_alt"],
                        border_width=1,
                        border_color=UI_THEME["border"],
                    )
                else:
                    widget = tk.Text(
                        self,
                        height=4 if key == "observacao" else 3,
                        width=40,
                        bg=UI_THEME["surface_alt"],
                        fg=UI_THEME["text"],
                        relief="flat",
                        highlightthickness=1,
                        highlightbackground=UI_THEME["border"],
                        insertbackground=UI_THEME["text"],
                    )
                widget.insert("1.0", str(record.get(key) or ""))
            else:
                if self.use_ctk:
                    widget = ctk.CTkEntry(
                        self.form_parent,
                        width=430,
                        fg_color=UI_THEME["surface_alt"],
                        border_width=1,
                        border_color=UI_THEME["border"],
                        text_color=UI_THEME["text"],
                    )
                else:
                    widget = tk.Entry(
                        self,
                        width=45,
                        bg=UI_THEME["surface_alt"],
                        fg=UI_THEME["text"],
                        relief="flat",
                        highlightthickness=1,
                        highlightbackground=UI_THEME["border"],
                        insertbackground=UI_THEME["text"],
                    )
                value = record.get(key)
                if value is None:
                    value_str = ""
                elif key == "data_bo":
                    value_str = value.strftime("%d/%m/%Y")
                else:
                    value_str = str(value)
                widget.insert(0, value_str)

            if self.use_ctk:
                widget.grid(row=row, column=1, sticky="ew", padx=14, pady=4)
            else:
                widget.grid(row=row, column=1, sticky="ew", padx=12, pady=4)
            self.widgets[key] = widget
            if key == "data_bo":
                self._bind_data_bo_placeholder(widget)
                if not value_str:
                    self._set_data_bo_placeholder(widget)
            row += 1

        if self.use_ctk:
            ctk.CTkLabel(
                self.form_parent,
                text="Alerta (min)",
                font=("Segoe UI", 12),
                text_color=UI_THEME["muted"],
            ).grid(row=row, column=0, sticky="w", padx=14, pady=8)
            ctk.CTkComboBox(
                self.form_parent,
                textvariable=self.intervalo_var,
                state="readonly",
                values=[str(minutes) for _, minutes in ALERT_INTERVAL_OPTIONS],
                width=180,
            ).grid(row=row, column=1, sticky="w", padx=14, pady=8)
        else:
            tk.Label(
                self,
                text="Alerta (min)",
                bg=UI_THEME["surface"],
                fg=UI_THEME["muted"],
                font=("Segoe UI", 10),
            ).grid(row=row, column=0, sticky="w", padx=12, pady=8)
            ttk.Combobox(
                self,
                textvariable=self.intervalo_var,
                state="readonly",
                values=[str(minutes) for _, minutes in ALERT_INTERVAL_OPTIONS],
                style="AFIS.TCombobox",
            ).grid(
                row=row, column=1, sticky="w", padx=12, pady=8
            )
        row += 1

        if self.use_ctk:
            actions = ctk.CTkFrame(self.form_parent, fg_color="transparent")
            actions.grid(row=row, column=0, columnspan=2, sticky="ew", padx=14, pady=(8, 14))
            ctk.CTkButton(
                actions,
                text="Salvar",
                command=self.save,
                fg_color=UI_THEME["success"],
                hover_color=UI_THEME["success_hover"],
                text_color=UI_THEME["white"],
                font=BUTTON_FONT_BOLD,
                corner_radius=8,
                height=34,
                width=130,
            ).pack(side="left")
            _build_button(actions, "Cancelar", self.destroy, "neutral", use_ctk=True, width=130).pack(side="left", padx=(8, 0))
        else:
            actions = tk.Frame(self, bg=UI_THEME["surface"])
            actions.grid(row=row, column=0, columnspan=2, sticky="ew", padx=12, pady=12)
            _build_button(actions, "Salvar", self.save, "success").pack(side="left")
            _build_button(actions, "Cancelar", self.destroy, "neutral").pack(side="left", padx=(8, 0))

        self.form_parent.columnconfigure(1, weight=1)
        _center_toplevel_on_parent(self, parent)
        self.transient(parent)
        self.grab_set()

    def _collect(self):
        """Coleta e normaliza os valores atuais do formulario de edicao."""
        data = {}
        for key in EDITABLE_FIELDS:
            widget = self.widgets[key]
            if isinstance(widget, tk.Text) or (ctk is not None and isinstance(widget, ctk.CTkTextbox)):
                raw = widget.get("1.0", tk.END).strip()
            else:
                if key == "data_bo" and self.data_bo_placeholder_active:
                    raw = ""
                else:
                    raw = widget.get().strip()
            data[key] = _normalize_user_text(raw, key)
        data_solic = self.record.get("data_solic")
        hora_solic = self.record.get("hora_solic")
        data["data_solic"] = data_solic.strftime("%d/%m/%Y") if data_solic else ""
        data["hora_solic"] = str(hora_solic)[:5] if hora_solic else ""
        return data

    def _set_entry_text_color(self, widget, color):
        """Ajusta cor do texto em Entry Tk/CTk."""
        if ctk is not None and isinstance(widget, ctk.CTkEntry):
            widget.configure(text_color=color)
        else:
            widget.configure(fg=color)

    def _bind_data_bo_placeholder(self, widget):
        """Vincula eventos de placeholder ao campo data BO."""
        widget.bind("<FocusIn>", lambda _e: self._on_data_bo_focus_in(widget))
        widget.bind("<FocusOut>", lambda _e: self._on_data_bo_focus_out(widget))
        widget.bind("<KeyPress>", lambda e: self._on_data_bo_key_press(widget, e))

    def _set_data_bo_placeholder(self, widget):
        """Define texto de placeholder para data BO."""
        widget.delete(0, tk.END)
        widget.insert(0, "dd/mm/aaaa")
        self._set_entry_text_color(widget, UI_THEME["placeholder"])
        self.data_bo_placeholder_active = True

    def _on_data_bo_focus_in(self, widget):
        """Mantem cursor no inicio quando placeholder estiver ativo."""
        if self.data_bo_placeholder_active:
            widget.icursor(0)

    def _on_data_bo_focus_out(self, widget):
        """Reaplica placeholder ao perder foco com campo vazio."""
        if not widget.get().strip():
            self._set_data_bo_placeholder(widget)

    def _on_data_bo_key_press(self, widget, event):
        """Remove placeholder ao iniciar digitacao no campo data BO."""
        if self.data_bo_placeholder_active and (event.char and event.char.isprintable()):
            widget.delete(0, tk.END)
            self._set_entry_text_color(widget, UI_THEME["text"])
            self.data_bo_placeholder_active = False

    def save(self):
        """Valida e persiste alteracoes do talao em edicao."""
        data = self._collect()

        try:
            normalized, missing = self.talao_service.prepare_update_talao(data)
        except ValueError as exc:
            messagebox.showwarning("Validacao", str(exc))
            return

        if missing:
            messagebox.showwarning("Validação", "Campos obrigatórios:\n- " + "\n- ".join(missing))
            return

        if normalized.get("status") == STATUS_FINALIZADO:
            confirmar_envio = messagebox.askyesno(
                "Confirmação obrigatória",
                self.alerta_service.build_final_boletim_confirmation_question(),
                parent=self,
            )
            if not confirmar_envio:
                normalized["status"] = STATUS_MONITORADO
                status_widget = self.widgets.get("status")
                if status_widget is not None and hasattr(status_widget, "set"):
                    status_widget.set(STATUS_MONITORADO)
                messagebox.showinfo(
                    "Status mantido",
                    "O talão permanecerá como MONITORADO porque o boletim finalizado ainda não foi enviado.",
                    parent=self,
                )

        try:
            self.repo.update_talao(
                self.talao_id,
                normalized,
                int(self.intervalo_var.get()),
                expected_updated_at=self.record.get("atualizado_em"),
            )
            self.on_saved()
            self.destroy()
        except ConcurrencyError as exc:
            messagebox.showwarning("Conflito de edição", str(exc))
            self.on_saved()
            self.destroy()
        except Exception:
            logger.exception("Falha ao atualizar talão %s", self.talao_id)
            messagebox.showerror("Erro", "Falha ao atualizar talão. Verifique os dados e tente novamente.")


class RelatorioPeriodoWindow(tk.Toplevel):
    """Janela modal para gerar relatorios por periodo."""

    def __init__(self, parent, repo: TalaoRepository):
        super().__init__(parent)
        self.repo = repo
        self.title("Relatórios por Período")
        self.geometry("210x180")
        self.minsize(210, 180)
        self.resizable(True, True)
        self.date_placeholder_active = {"inicio": False, "fim": False}
        self.use_ctk = ctk is not None

        _apply_toplevel_theme(self)

        if self.use_ctk:
            container = ctk.CTkFrame(
                self,
                fg_color=UI_THEME["surface"],
                corner_radius=12,
                border_width=1,
                border_color=UI_THEME["border"],
            )
            container.pack(fill="both", expand=True, padx=14, pady=14)

            ctk.CTkLabel(
                container,
                text="Relatório de Talões",
                font=("Segoe UI", 16, "bold"),
                text_color=UI_THEME["text"],
            ).grid(row=0, column=0, columnspan=2, sticky="w", padx=14, pady=(14, 10))

            ctk.CTkLabel(container, text="Data início", text_color=UI_THEME["muted"]).grid(
                row=1, column=0, sticky="w", padx=14, pady=4
            )
            self.data_inicio_entry = ctk.CTkEntry(container, width=180, fg_color=UI_THEME["surface_alt"])
            self.data_inicio_entry.grid(row=1, column=1, sticky="ew", padx=14, pady=4)

            ctk.CTkLabel(container, text="Data fim", text_color=UI_THEME["muted"]).grid(
                row=2, column=0, sticky="w", padx=14, pady=4
            )
            self.data_fim_entry = ctk.CTkEntry(container, width=180, fg_color=UI_THEME["surface_alt"])
            self.data_fim_entry.grid(row=2, column=1, sticky="ew", padx=14, pady=4)

            actions = ctk.CTkFrame(container, fg_color="transparent")
            actions.grid(row=3, column=0, columnspan=2, sticky="ew", padx=14, pady=(14, 14))
            _build_button(
                actions,
                "Excel",
                self.gerar_modelo_xlsx,
                "success",
                use_ctk=True,
                width=170,
            ).pack(side="left")
            ctk.CTkButton(
                actions,
                text="CSV",
                command=self.gerar_csv,
                fg_color=UI_THEME["warning"],
                hover_color=UI_THEME["warning_hover"],
                text_color=UI_THEME["white"],
                font=BUTTON_FONT_BOLD,
                width=120,
            ).pack(side="left", padx=(8, 0))
            _build_button(actions, "Cancelar", self.destroy, "neutral", use_ctk=True, width=120).pack(side="left", padx=(8, 0))

            container.columnconfigure(1, weight=1)
        else:
            frame = tk.Frame(self, padx=12, pady=12, bg=UI_THEME["surface"])
            frame.pack(fill="both", expand=True)
            frame.columnconfigure(1, weight=1)

            tk.Label(
                frame,
                text="Relatório de Talões",
                font=("Segoe UI", 12, "bold"),
                bg=UI_THEME["surface"],
                fg=UI_THEME["text"],
            ).grid(
                row=0, column=0, columnspan=2, sticky="w", pady=(0, 10)
            )
            tk.Label(frame, text="Data início", bg=UI_THEME["surface"], fg=UI_THEME["muted"]).grid(row=1, column=0, sticky="w", pady=4)
            self.data_inicio_entry = tk.Entry(
                frame,
                width=24,
                bg=UI_THEME["surface_alt"],
                fg=UI_THEME["text"],
                relief="flat",
                highlightthickness=1,
                highlightbackground=UI_THEME["border"],
                insertbackground=UI_THEME["text"],
            )
            self.data_inicio_entry.grid(row=1, column=1, sticky="ew", pady=4)

            tk.Label(frame, text="Data fim", bg=UI_THEME["surface"], fg=UI_THEME["muted"]).grid(row=2, column=0, sticky="w", pady=4)
            self.data_fim_entry = tk.Entry(
                frame,
                width=24,
                bg=UI_THEME["surface_alt"],
                fg=UI_THEME["text"],
                relief="flat",
                highlightthickness=1,
                highlightbackground=UI_THEME["border"],
                insertbackground=UI_THEME["text"],
            )
            self.data_fim_entry.grid(row=2, column=1, sticky="ew", pady=4)

            actions = tk.Frame(frame, bg=UI_THEME["surface"])
            actions.grid(row=3, column=0, columnspan=2, sticky="ew", pady=(14, 0))
            _build_button(actions, "Excel", self.gerar_modelo_xlsx, "success").pack(side="left")
            _build_button(actions, "CSV", self.gerar_csv, "warning").pack(side="left", padx=(8, 0))
            _build_button(actions, "Cancelar", self.destroy, "neutral").pack(side="left", padx=(8, 0))

        self._bind_date_placeholder(self.data_inicio_entry, "inicio")
        self._bind_date_placeholder(self.data_fim_entry, "fim")
        self._set_date_placeholder(self.data_inicio_entry, "inicio")
        self._set_date_placeholder(self.data_fim_entry, "fim")

        _center_toplevel_on_parent(self, parent)
        self.transient(parent)
        self.grab_set()

    def _parse_periodo(self):
        """Converte e valida datas de inicio/fim informadas no formulario."""
        data_inicio_txt = self._get_date_value(self.data_inicio_entry, "inicio")
        data_fim_txt = self._get_date_value(self.data_fim_entry, "fim")
        try:
            data_inicio = datetime.strptime(data_inicio_txt, "%d/%m/%Y").date()
            data_fim = datetime.strptime(data_fim_txt, "%d/%m/%Y").date()
        except ValueError as exc:
            raise ValueError("Datas inválidas. Use o formato dd/mm/aaaa.") from exc
        if data_inicio > data_fim:
            raise ValueError("Data início não pode ser maior que data fim.")
        return data_inicio, data_fim

    def _set_entry_text_color(self, widget, color):
        """Ajusta cor do texto em Entry Tk/CTk da janela de relatorio."""
        if ctk is not None and isinstance(widget, ctk.CTkEntry):
            widget.configure(text_color=color)
        else:
            widget.configure(fg=color)

    def _bind_date_placeholder(self, widget, key):
        """Vincula eventos de placeholder aos campos de periodo."""
        widget.bind("<FocusIn>", lambda _e: self._on_date_focus_in(widget, key))
        widget.bind("<FocusOut>", lambda _e: self._on_date_focus_out(widget, key))
        widget.bind("<KeyPress>", lambda e: self._on_date_key_press(widget, key, e))

    def _set_date_placeholder(self, widget, key):
        """Aplica placeholder de data em campo de periodo."""
        widget.delete(0, tk.END)
        widget.insert(0, "dd/mm/aaaa")
        self._set_entry_text_color(widget, UI_THEME["placeholder"])
        self.date_placeholder_active[key] = True

    def _on_date_focus_in(self, widget, key):
        """Mantem cursor no inicio quando placeholder de data estiver ativo."""
        if self.date_placeholder_active[key]:
            widget.icursor(0)

    def _on_date_focus_out(self, widget, key):
        """Reaplica placeholder de data quando campo ficar vazio."""
        if not widget.get().strip():
            self._set_date_placeholder(widget, key)

    def _on_date_key_press(self, widget, key, event):
        """Remove placeholder de data ao iniciar digitacao."""
        if self.date_placeholder_active[key] and (event.char and event.char.isprintable()):
            widget.delete(0, tk.END)
            self._set_entry_text_color(widget, UI_THEME["text"])
            self.date_placeholder_active[key] = False

    def _get_date_value(self, widget, key):
        """Retorna valor real do campo ignorando placeholder ativo."""
        if self.date_placeholder_active[key]:
            return ""
        return widget.get().strip()

    def _load_report_rows(self):
        """Carrega linhas de relatorio para o periodo informado."""
        try:
            data_inicio, data_fim = self._parse_periodo()
        except ValueError as exc:
            messagebox.showwarning("Validação", str(exc))
            return None

        try:
            columns, rows = self.repo.list_taloes_by_period(data_inicio, data_fim)
        except Exception:
            logger.exception("Falha ao consultar talões para relatório")
            messagebox.showerror("Erro", "Falha ao consultar dados para o relatório.")
            return None
        return data_inicio, data_fim, columns, rows

    def _resolve_modelo_path(self):
        """Retorna caminho absoluto do template XLSX de relatorio."""
        return Path(__file__).resolve().parent.parent / "assets" / "modelo.xlsx"

    def _format_excel_date(self, value):
        """Converte valores de data para formato de exibicao no Excel."""
        if value is None:
            return ""
        if isinstance(value, datetime):
            return value.strftime("%d/%m/%Y")
        if isinstance(value, date):
            return value.strftime("%d/%m/%Y")
        return str(value)

    def gerar_csv(self):
        """Exporta relatorio de periodo para arquivo CSV."""
        loaded = self._load_report_rows()
        if loaded is None:
            return
        data_inicio, data_fim, columns, rows = loaded

        nome_base = f"relatorio_taloes_{data_inicio.strftime('%Y%m%d')}_{data_fim.strftime('%Y%m%d')}.csv"
        path = filedialog.asksaveasfilename(
            title="Salvar relatório CSV",
            defaultextension=".csv",
            initialfile=nome_base,
            filetypes=[("CSV", "*.csv"), ("Todos os arquivos", "*.*")],
        )
        if not path:
            return

        try:
            with open(path, "w", encoding="utf-8-sig", newline="") as csv_file:
                writer = csv.writer(csv_file, delimiter=";")
                writer.writerow(columns)
                for row in rows:
                    writer.writerow(list(row))
        except Exception:
            logger.exception("Falha ao gravar relatório CSV em %s", path)
            messagebox.showerror("Erro", "Falha ao gravar arquivo CSV.")
            return

        messagebox.showinfo("Relatório", f"Relatório gerado com sucesso.\nRegistros exportados: {len(rows)}")
        self.destroy()

    def gerar_modelo_xlsx(self):
        """Exporta relatorio para XLSX usando template institucional."""
        if load_workbook is None:
            messagebox.showerror(
                "Dependência ausente",
                "A biblioteca openpyxl não está instalada.\nInstale para habilitar a geração de XLSX pelo modelo.",
            )
            return

        loaded = self._load_report_rows()
        if loaded is None:
            return
        data_inicio, data_fim, columns, rows = loaded

        modelo_path = self._resolve_modelo_path()
        if not modelo_path.exists():
            messagebox.showerror("Erro", f"Modelo não encontrado:\n{modelo_path}")
            return

        nome_base = f"relatorio_taloes_modelo_{data_inicio.strftime('%Y%m%d')}_{data_fim.strftime('%Y%m%d')}.xlsx"
        path = filedialog.asksaveasfilename(
            title="Salvar relatório XLSX (modelo)",
            defaultextension=".xlsx",
            initialfile=nome_base,
            filetypes=[("Excel", "*.xlsx"), ("Todos os arquivos", "*.*")],
        )
        if not path:
            return

        try:
            wb = load_workbook(modelo_path)
            ws = wb.active
            col_idx = {name: idx for idx, name in enumerate(columns)}

            max_row = max(ws.max_row, 7)
            for row_idx in range(7, max_row + 1):
                for col in range(1, 9):
                    ws.cell(row=row_idx, column=col, value=None)

            row_excel = 7
            for row in rows:
                values = list(row)
                ano = values[col_idx["ano"]]
                talao = values[col_idx["talao"]]
                ws.cell(row=row_excel, column=1, value=self._format_excel_date(values[col_idx["data_solic"]]))
                ws.cell(row=row_excel, column=2, value=format_talao(ano, talao))
                ws.cell(row=row_excel, column=3, value=self._format_excel_date(values[col_idx["data_bo"]]))
                ws.cell(row=row_excel, column=4, value=values[col_idx["boletim"]] or "")
                ws.cell(row=row_excel, column=5, value=values[col_idx["delegacia"]] or "")
                ws.cell(row=row_excel, column=6, value=values[col_idx["natureza"]] or "")
                ws.cell(row=row_excel, column=7, value=values[col_idx["vitimas"]] or "")
                ws.cell(row=row_excel, column=8, value=values[col_idx["equipe"]] or "")
                row_excel += 1

            wb.save(path)
        except Exception:
            logger.exception("Falha ao gerar XLSX de relatório com modelo em %s", path)
            messagebox.showerror("Erro", "Falha ao gerar arquivo XLSX pelo modelo.")
            return

        messagebox.showinfo("Relatório", f"Relatório XLSX gerado com sucesso.\nRegistros exportados: {len(rows)}")
        self.destroy()


class BackupAnoWindow(tk.Toplevel):
    """Janela modal para gerar backup SQL por ano de referencia."""

    def __init__(self, parent, repo: TalaoRepository):
        super().__init__(parent)
        self.repo = repo
        self.title("Backup por Ano")
        self.geometry("210x130")
        self.minsize(210, 130)
        self.resizable(True, True)
        self.ano_var = tk.StringVar(value=str(datetime.now().year - 1))
        self.use_ctk = ctk is not None

        _apply_toplevel_theme(self)

        if self.use_ctk:
            container = ctk.CTkFrame(
                self,
                fg_color=UI_THEME["surface"],
                corner_radius=12,
                border_width=1,
                border_color=UI_THEME["border"],
            )
            container.pack(fill="both", expand=True, padx=14, pady=14)

            ctk.CTkLabel(
                container,
                text="Backup dos Talões",
                font=("Segoe UI", 16, "bold"),
                text_color=UI_THEME["text"],
            ).grid(row=0, column=0, sticky="w", padx=14, pady=(14, 8))
            ctk.CTkLabel(
                container,
                text="Ano de referência",
                text_color=UI_THEME["muted"],
            ).grid(row=1, column=0, sticky="w", padx=14, pady=(0, 6))
            ctk.CTkEntry(
                container,
                textvariable=self.ano_var,
                width=120,
                fg_color=UI_THEME["surface_alt"],
                border_width=1,
                border_color=UI_THEME["border"],
            ).grid(row=1, column=1, sticky="ew", padx=14, pady=(0, 6))

            actions = ctk.CTkFrame(container, fg_color="transparent")
            actions.grid(row=2, column=0, columnspan=2, sticky="ew", padx=14, pady=(6, 12))
            ctk.CTkButton(
                actions,
                text="Backup SQL",
                command=self.gerar_backup,
                fg_color=UI_THEME["danger"],
                hover_color=UI_THEME["danger_hover"],
                text_color=UI_THEME["white"],
                font=BUTTON_FONT_BOLD,
                width=160,
            ).pack(side="left")
            _build_button(actions, "Cancelar", self.destroy, "neutral", use_ctk=True, width=120).pack(side="left", padx=(8, 0))
            container.columnconfigure(1, weight=1)
        else:
            frame = tk.Frame(self, padx=12, pady=12, bg=UI_THEME["surface"])
            frame.pack(fill="both", expand=True)
            frame.columnconfigure(1, weight=1)
            tk.Label(
                frame,
                text="Backup dos Talões",
                font=("Segoe UI", 12, "bold"),
                bg=UI_THEME["surface"],
                fg=UI_THEME["text"],
            ).grid(row=0, column=0, sticky="w", pady=(0, 8))
            tk.Label(frame, text="Ano de referência", bg=UI_THEME["surface"], fg=UI_THEME["muted"]).grid(
                row=1, column=0, sticky="w", pady=(0, 6)
            )
            tk.Entry(
                frame,
                textvariable=self.ano_var,
                width=8,
                bg=UI_THEME["surface_alt"],
                fg=UI_THEME["text"],
                relief="flat",
                highlightthickness=1,
                highlightbackground=UI_THEME["border"],
                insertbackground=UI_THEME["text"],
            ).grid(row=1, column=1, sticky="ew", pady=(0, 6))
            actions = tk.Frame(frame, bg=UI_THEME["surface"])
            actions.grid(row=2, column=0, columnspan=2, sticky="ew", pady=(6, 0))
            _build_button(actions, "Backup SQL", self.gerar_backup, "danger").pack(side="left")
            _build_button(actions, "Cancelar", self.destroy, "neutral").pack(side="left", padx=(8, 0))

        _center_toplevel_on_parent(self, parent)
        self.transient(parent)
        self.grab_set()

    def _sql_literal(self, value):
        """Converte valor Python para literal SQL seguro para script."""
        if value is None:
            return "NULL"
        if isinstance(value, bool):
            return "1" if value else "0"
        if isinstance(value, datetime):
            return f"'{value.strftime('%Y-%m-%d %H:%M:%S.%f')[:-3]}'"
        if isinstance(value, date):
            return f"'{value.strftime('%Y-%m-%d')}'"
        if isinstance(value, time):
            return f"'{value.strftime('%H:%M:%S')}'"
        if isinstance(value, (int, float)):
            return str(value)
        text = str(value).replace("'", "''")
        return f"N'{text}'"

    def _build_insert_block(self, table_name, columns, rows):
        """Monta bloco de INSERTs para uma tabela e conjunto de linhas."""
        if not rows:
            return [f"-- Nenhum registro para {table_name}."]
        cols_sql = ", ".join(f"[{col}]" for col in columns)
        lines = [f"SET IDENTITY_INSERT {table_name} ON;"]
        for row in rows:
            values_sql = ", ".join(self._sql_literal(v) for v in row)
            lines.append(f"INSERT INTO {table_name} ({cols_sql}) VALUES ({values_sql});")
        lines.append(f"SET IDENTITY_INSERT {table_name} OFF;")
        return lines

    def gerar_backup(self):
        """Gera arquivo SQL de backup contendo dados de um ano."""
        ano_txt = self.ano_var.get().strip()
        try:
            ano = int(ano_txt)
        except ValueError:
            messagebox.showwarning("Validação", "Informe um ano válido com 4 dígitos.")
            return
        if ano < 1900 or ano > 9999:
            messagebox.showwarning("Validação", "Informe um ano válido entre 1900 e 9999.")
            return

        try:
            taloes_cols, taloes_rows = self.repo.list_taloes_by_year(ano)
            mon_cols, mon_rows = self.repo.list_monitoramento_by_year(ano)
        except Exception:
            logger.exception("Falha ao coletar dados para backup do ano %s", ano)
            messagebox.showerror("Erro", "Falha ao consultar dados para backup.")
            return

        nome_base = f"backup_afis_{ano}.sql"
        path = filedialog.asksaveasfilename(
            title="Salvar backup SQL",
            defaultextension=".sql",
            initialfile=nome_base,
            filetypes=[("SQL", "*.sql"), ("Todos os arquivos", "*.*")],
        )
        if not path:
            return

        lines = [
            f"-- Backup AFIS ano {ano}",
            "SET NOCOUNT ON;",
            "BEGIN TRANSACTION;",
            "BEGIN TRY",
            "",
            f"-- Tabela dbo.taloes ({len(taloes_rows)} registros)",
        ]
        lines.extend(self._build_insert_block("dbo.taloes", taloes_cols, taloes_rows))
        lines.append("")
        lines.append(f"-- Tabela dbo.monitoramento ({len(mon_rows)} registros)")
        lines.extend(self._build_insert_block("dbo.monitoramento", mon_cols, mon_rows))
        lines.extend(
            [
                "",
                "COMMIT TRANSACTION;",
                "END TRY",
                "BEGIN CATCH",
                "    IF @@TRANCOUNT > 0 ROLLBACK TRANSACTION;",
                "    THROW;",
                "END CATCH;",
                "",
            ]
        )

        try:
            with open(path, "w", encoding="utf-8-sig", newline="") as f:
                f.write("\n".join(lines))
        except Exception:
            logger.exception("Falha ao gravar backup SQL em %s", path)
            messagebox.showerror("Erro", "Falha ao gravar arquivo de backup.")
            return

        messagebox.showinfo(
            "Backup concluído",
            f"Arquivo gerado com sucesso.\nTalões: {len(taloes_rows)}\nMonitoramento: {len(mon_rows)}",
        )
        self.destroy()


class AFISDashboard:
    """Tela principal do sistema AFIS com operacoes de cadastro e monitoramento."""

    ALERT_POLL_MS = 30000
    AUTO_REFRESH_MS = 60000

    def __init__(self, root, repo: TalaoRepository):
        """Inicializa estado da tela principal e agenda rotinas automaticas."""
        self.root = root
        self.repo = repo
        self.talao_service = TalaoService()
        self.alerta_service = AlertaService()

        self.root.title("Registro AFIS - CECOP")
        self.root.geometry("1080x760")
        self._apply_theme()

        self.intervalo_map = dict(ALERT_INTERVAL_OPTIONS)

        self.widgets = {}
        self.watermark_image = None
        self.watermark_label = None
        self.help_icon_image = None
        self.whatsapp_icon_image = None
        self.data_bo_placeholder_active = False
        self.proximo_talao_var = tk.StringVar(value="-")
        self.alerta_var = tk.StringVar(value=DEFAULT_ALERT_INTERVAL_LABEL)

        self._setup_watermark()
        self._build_layout()
        self._set_defaults()
        self.refresh_tree()
        self.root.after(self.AUTO_REFRESH_MS, self._auto_refresh)
        self.root.after(self.ALERT_POLL_MS, self.processar_alertas)

    def _apply_theme(self):
        """Configura estilos visuais globais da interface principal."""
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("AFIS.TLabelframe", background=UI_THEME["surface"], borderwidth=1)
        style.configure("AFIS.TLabelframe.Label", background=UI_THEME["surface"], foreground=UI_THEME["muted"], font=("Segoe UI", 10, "bold"))
        style.configure("AFIS.Treeview", background=UI_THEME["surface"], foreground=UI_THEME["text"], fieldbackground=UI_THEME["surface"], rowheight=28)
        style.configure("AFIS.Treeview.Heading", background=UI_THEME["surface_alt"], foreground=UI_THEME["text"], font=("Segoe UI", 9, "bold"))
        style.map("AFIS.Treeview.Heading", background=[("active", UI_THEME["surface_hover"])])
        style.configure("AFIS.TCombobox", padding=4)
        if ctk is not None and isinstance(self.root, ctk.CTk):
            self.root.configure(fg_color=UI_THEME["bg"])
        else:
            self.root.configure(bg=UI_THEME["bg"])

    def _build_button(self, parent, text, command, variant="neutral"):
        """Cria botao padronizado no contexto do dashboard."""
        return _build_button(parent, text, command, variant=variant, use_ctk=ctk is not None)

    def _build_layout(self):
        """Monta layout completo do dashboard principal."""
        header = tk.Frame(self.root, bg=UI_THEME["bg"])
        header.pack(fill="x", padx=12, pady=(14, 10))
        header.grid_columnconfigure(0, weight=1)
        header.grid_columnconfigure(1, weight=0)
        header.grid_columnconfigure(2, weight=1)

        titulo = tk.Label(
            header,
            text="Registro Digital de Talões AFIS",
            font=("Segoe UI", 20, "bold"),
            bg=UI_THEME["bg"],
            fg=UI_THEME["white"],
        )
        titulo.grid(row=0, column=1)

        self.help_icon_image = self._load_help_icon()
        if self.help_icon_image is not None:
            ajuda_btn = tk.Button(
                header,
                image=self.help_icon_image,
                command=self.abrir_manual_usuario,
                bg=UI_THEME["bg"],
                activebackground=UI_THEME["bg"],
                borderwidth=0,
                highlightthickness=0,
                cursor="hand2",
            )
        else:
            ajuda_btn = tk.Button(
                header,
                text="?",
                command=self.abrir_manual_usuario,
                font=("Segoe UI", 12, "bold"),
                bg=UI_THEME["neutral"],
                fg=UI_THEME["text"],
                activebackground=UI_THEME["neutral_hover"],
                borderwidth=0,
                cursor="hand2",
                width=3,
            )
        ajuda_btn.grid(row=0, column=2, sticky="e")

        form = ttk.LabelFrame(self.root, text="Abertura de Talão", style="AFIS.TLabelframe", padding=12)
        form.pack(fill="x", padx=12, pady=6)

        info = tk.Frame(form, bg=UI_THEME["surface"])
        info.grid(row=0, column=0, columnspan=4, sticky="ew", pady=(0, 6))
        tk.Label(
            info,
            text="Próximo Talão (ano atual):",
            font=("Segoe UI", 10, "bold"),
            bg=UI_THEME["surface"],
            fg=UI_THEME["muted"],
        ).pack(side="left")
        tk.Label(
            info,
            textvariable=self.proximo_talao_var,
            fg=UI_THEME["primary"],
            bg=UI_THEME["surface"],
            font=("Segoe UI", 10, "bold"),
        ).pack(side="left", padx=6)

        layout_fields = [
            "delegacia",
            "autoridade",
            "solicitante",
            "endereco",
            "boletim",
            "natureza",
            "data_bo",
            "vitimas",
            "equipe",
            "operador",
            "observacao",
        ]

        row = 1
        col = 0
        for key in layout_fields:
            label_text = FIELD_LABELS[key]
            if key == "data_bo":
                label_text = "Data BO"
            tk.Label(
                form,
                text=label_text,
                bg=UI_THEME["surface"],
                fg=UI_THEME["muted"],
                font=("Segoe UI", 9),
            ).grid(row=row, column=col, sticky="w", pady=4, padx=4)

            if key == "status":
                widget = ttk.Combobox(form, values=STATUS_OPCOES, state="readonly", width=28, style="AFIS.TCombobox")
                widget.set(STATUS_MONITORADO)
            elif key == "observacao":
                widget = tk.Text(
                    form,
                    width=35,
                    height=3,
                    bg=UI_THEME["surface_alt"],
                    fg=UI_THEME["text"],
                    relief="flat",
                    highlightthickness=1,
                    highlightbackground=UI_THEME["border"],
                    insertbackground=UI_THEME["text"],
                )
            else:
                widget = tk.Entry(
                    form,
                    width=32,
                    bg=UI_THEME["surface_alt"],
                    fg=UI_THEME["text"],
                    relief="flat",
                    highlightthickness=1,
                    highlightbackground=UI_THEME["border"],
                    insertbackground=UI_THEME["text"],
                )

            widget.grid(row=row, column=col + 1, sticky="ew", pady=4, padx=4)
            self.widgets[key] = widget
            if key == "data_bo":
                self._bind_data_bo_placeholder(widget)

            if col == 2:
                col = 0
                row += 1
            else:
                col = 2

        row += 1
        tk.Label(
            form,
            text="Alerta (min)",
            bg=UI_THEME["surface"],
            fg=UI_THEME["muted"],
            font=("Segoe UI", 9),
        ).grid(row=row, column=0, sticky="w", padx=4, pady=8)
        combo_alerta = ttk.Combobox(
            form,
            textvariable=self.alerta_var,
            state="readonly",
            values=list(self.intervalo_map.keys()),
            style="AFIS.TCombobox",
        )
        combo_alerta.grid(row=row, column=1, sticky="w", padx=4, pady=8)

        botoes = tk.Frame(form, bg=UI_THEME["surface"])
        botoes.grid(row=row, column=2, columnspan=2, sticky="ew", padx=4, pady=8)
        self._build_button(botoes, "Salvar", self.criar_talao, "success").pack(side="left", padx=4)
        self._build_button(botoes, "Editar", self.editar_selecionado, "primary").pack(side="left", padx=4)
        self._build_button(botoes, "Atualizar", self.refresh_tree, "neutral").pack(side="left", padx=4)
        self._build_button(botoes, "Limpar", self._set_defaults, "neutral").pack(side="left", padx=4)
        self._build_button(botoes, "Relatórios", self.abrir_relatorios, "warning").pack(side="left", padx=4)
        self._build_button(botoes, "Backup", self.abrir_backup, "danger").pack(side="left", padx=4)
        self.whatsapp_icon_image = self._load_whatsapp_icon()
        if self.whatsapp_icon_image is not None:
            tk.Button(
                botoes,
                image=self.whatsapp_icon_image,
                command=self.gerar_mensagem_whatsapp_selecionado,
                bg=UI_THEME["surface"],
                activebackground=UI_THEME["surface_hover"],
                relief="flat",
                borderwidth=1,
                highlightthickness=1,
                highlightbackground=UI_THEME["border"],
                cursor="hand2",
                padx=6,
                pady=4,
            ).pack(side="left", padx=4)
        else:
            self._build_button(
                botoes,
                "WhatsApp",
                self.gerar_mensagem_whatsapp_selecionado,
                "neutral",
            ).pack(side="left", padx=4)

        for col_idx in (1, 3):
            form.grid_columnconfigure(col_idx, weight=1)

        list_frame = ttk.LabelFrame(self.root, text="Talões visíveis", style="AFIS.TLabelframe", padding=12)
        list_frame.pack(fill="both", expand=True, padx=12, pady=8)
        
        cols = ("talao", "boletim", "delegacia", "natureza", "status")
        self.tree = ttk.Treeview(list_frame, columns=cols, show="headings", height=15, style="AFIS.Treeview")
        self.tree.heading("talao", text="Talão")
        self.tree.heading("boletim", text="Boletim")
        self.tree.heading("delegacia", text="Delegacia")
        self.tree.heading("natureza", text="Natureza")
        self.tree.heading("status", text="Status")

        self.tree.column("talao", width=90, anchor="center")
        self.tree.column("boletim", width=120, anchor="center")
        self.tree.column("delegacia", width=220)
        self.tree.column("natureza", width=240)
        self.tree.column("status", width=120, anchor="center")

        self.tree.tag_configure(
            STATUS_MONITORADO,
            background=UI_THEME["status_monitorado_bg"],
            foreground=UI_THEME["status_monitorado_fg"],
        )
        self.tree.tag_configure(
            STATUS_FINALIZADO,
            background=UI_THEME["status_finalizado_bg"],
            foreground=UI_THEME["status_finalizado_fg"],
        )
        self.tree.tag_configure(
            STATUS_CANCELADO,
            background=UI_THEME["status_cancelado_bg"],
            foreground=UI_THEME["status_cancelado_fg"],
        )

        self.tree.pack(fill="both", expand=True)

    def _set_defaults(self):
        """Restaura valores padrao dos campos de abertura."""
        defaults = {
            "status": STATUS_MONITORADO,
            "data_bo": "",
            "delegacia": "",
            "autoridade": "",
            "solicitante": "",
            "endereco": "",
            "boletim": "",
            "natureza": "",
            "vitimas": "",
            "equipe": "",
            "operador": "",
            "observacao": "",
        }

        for key, widget in self.widgets.items():
            value = defaults.get(key, "")
            if isinstance(widget, tk.Text):
                widget.delete("1.0", tk.END)
                widget.insert("1.0", value)
            elif isinstance(widget, ttk.Combobox):
                widget.set(value)
            else:
                widget.delete(0, tk.END)
                widget.insert(0, value)

        data_bo_widget = self.widgets.get("data_bo")
        if data_bo_widget is not None:
            self._set_data_bo_placeholder(data_bo_widget)
        self._refresh_proximo_talao()

    def _resolve_asset_path(self, path_value):
        """Resolve caminho de asset relativo ao diretorio raiz do projeto."""
        if not path_value:
            return None
        candidate = Path(path_value).expanduser()
        if candidate.is_absolute():
            return candidate
        project_root = Path(__file__).resolve().parent.parent
        return (project_root / candidate).resolve()

    def _load_watermark_image(self):
        """Carrega e redimensiona imagem de marca d'agua, quando configurada."""
        image_path = self._resolve_asset_path(get_env("APP_WATERMARK_IMAGE_PATH") or get_env("APP_HEADER_IMAGE_PATH"))
        if not image_path or not image_path.exists():
            return None
        try:
            image = tk.PhotoImage(file=str(image_path))
            width = image.width()
            height = image.height()
            if width <= 0 or height <= 0:
                return image

            width_ratio = (width + WATERMARK_MAX_WIDTH - 1) // WATERMARK_MAX_WIDTH
            height_ratio = (height + WATERMARK_MAX_HEIGHT - 1) // WATERMARK_MAX_HEIGHT
            scale = max(1, width_ratio, height_ratio)
            if scale > 1:
                image = image.subsample(scale, scale)
            return image
        except Exception:
            logger.warning("Falha ao carregar imagem da marca d'água em %s", image_path, exc_info=True)
            return None

    def _load_help_icon(self):
        """Carrega icone do botao de ajuda no topo da interface."""
        icon_path = self._resolve_asset_path("assets/help.png")
        if not icon_path or not icon_path.exists():
            return None
        try:
            image = tk.PhotoImage(file=str(icon_path))
            width = image.width()
            height = image.height()
            if width > 0 and height > 0:
                width_ratio = (width + HELP_ICON_SIZE_PX - 1) // HELP_ICON_SIZE_PX
                height_ratio = (height + HELP_ICON_SIZE_PX - 1) // HELP_ICON_SIZE_PX
                scale = max(1, width_ratio, height_ratio)
                if scale > 1:
                    image = image.subsample(scale, scale)
            return image
        except Exception:
            logger.warning("Falha ao carregar icone de ajuda em %s", icon_path, exc_info=True)
            return None

    def _load_whatsapp_icon(self):
        """Carrega icone do botao de mensagem WhatsApp."""
        icon_path = self._resolve_asset_path("assets/ZAP.png")
        if not icon_path or not icon_path.exists():
            return None
        try:
            image = tk.PhotoImage(file=str(icon_path))
            width = image.width()
            height = image.height()
            if width > 0 and height > 0:
                width_ratio = (width + WHATSAPP_ICON_SIZE_PX - 1) // WHATSAPP_ICON_SIZE_PX
                height_ratio = (height + WHATSAPP_ICON_SIZE_PX - 1) // WHATSAPP_ICON_SIZE_PX
                scale = max(1, width_ratio, height_ratio)
                if scale > 1:
                    image = image.subsample(scale, scale)
            return image
        except Exception:
            logger.warning("Falha ao carregar icone WhatsApp em %s", icon_path, exc_info=True)
            return None

    def _setup_watermark(self):
        """Posiciona marca d'agua no fundo da janela principal."""
        self.watermark_image = self._load_watermark_image()
        if self.watermark_image is None:
            return
        self.watermark_label = tk.Label(self.root, image=self.watermark_image, bg=UI_THEME["bg"], borderwidth=0, highlightthickness=0)
        self.watermark_label.place(relx=0.5, rely=0.56, anchor="center")
        self.watermark_label.lower()

    def _collect_form_data(self):
        """Coleta e normaliza dados do formulario de abertura."""
        data = {}
        for key, widget in self.widgets.items():
            if isinstance(widget, tk.Text):
                raw = widget.get("1.0", tk.END).strip()
            else:
                if key == "data_bo" and self.data_bo_placeholder_active:
                    raw = ""
                else:
                    raw = widget.get().strip()
            data[key] = _normalize_user_text(raw, key)
        return data

    def abrir_manual_usuario(self):
        """Abre o manual do usuario no navegador padrao do sistema."""
        candidates = [self._resolve_asset_path("MANUAL_USUARIO.html"), Path.cwd() / "MANUAL_USUARIO.html"]
        meipass = getattr(sys, "_MEIPASS", None)
        if meipass:
            candidates.insert(0, Path(meipass) / "MANUAL_USUARIO.html")
        candidates.insert(1, Path(sys.executable).resolve().parent / "MANUAL_USUARIO.html")

        manual_path = next((p for p in candidates if p and p.exists()), None)
        if manual_path is None:
            messagebox.showerror(
                "Manual não encontrado",
                "Não foi possível localizar o arquivo MANUAL_USUARIO.html.",
            )
            return

        try:
            webbrowser.open_new_tab(manual_path.resolve().as_uri())
        except Exception:
            logger.exception("Falha ao abrir manual do usuario em %s", manual_path)
            messagebox.showerror("Erro", "Não foi possível abrir o manual no navegador padrão.")

    def _bind_data_bo_placeholder(self, widget):
        """Vincula eventos de placeholder ao campo data BO principal."""
        widget.bind("<FocusIn>", lambda _e: self._on_data_bo_focus_in(widget))
        widget.bind("<FocusOut>", lambda _e: self._on_data_bo_focus_out(widget))
        widget.bind("<KeyPress>", lambda e: self._on_data_bo_key_press(widget, e))

    def _set_data_bo_placeholder(self, widget):
        """Aplica placeholder visual no campo data BO principal."""
        widget.delete(0, tk.END)
        widget.insert(0, "dd/mm/aaaa")
        widget.configure(fg=UI_THEME["placeholder"])
        self.data_bo_placeholder_active = True

    def _on_data_bo_focus_in(self, widget):
        """Mantem cursor no inicio quando placeholder estiver ativo."""
        if self.data_bo_placeholder_active:
            widget.icursor(0)

    def _on_data_bo_focus_out(self, widget):
        """Reaplica placeholder no campo data BO quando vazio."""
        if not widget.get().strip():
            self._set_data_bo_placeholder(widget)

    def _on_data_bo_key_press(self, widget, event):
        """Remove placeholder ao iniciar digitacao no campo data BO."""
        if self.data_bo_placeholder_active and (event.char and event.char.isprintable()):
            widget.delete(0, tk.END)
            widget.configure(fg=UI_THEME["text"])
            self.data_bo_placeholder_active = False

    def _refresh_proximo_talao(self):
        """Atualiza informacao de proximo numero de talao na interface."""
        try:
            ano = datetime.now().year
            numero = self.repo.get_next_talao(ano)
            self.proximo_talao_var.set(format_talao(ano, numero))
        except Exception:
            self.proximo_talao_var.set("indisponível")

    def _format_message_value(self, value):
        """Normaliza valor para exibicao no template de mensagem."""
        if value is None:
            return "-"
        text = str(value).strip()
        return text if text else "-"

    def _build_whatsapp_message(self, ano, talao, data):
        """Monta mensagem padrao para compartilhamento no WhatsApp."""
        lines = [
            "*NOVO TALÃO*\n",
            f"*TALÃO:{format_talao(ano, talao)}*",
            f"*DELEGACIA:* {self._format_message_value(data.get('delegacia'))}",
            f"*AUTORIDADE:* {self._format_message_value(data.get('autoridade'))}",
            f"*ENDERECO:* {self._format_message_value(data.get('endereco'))}",
            f"*BOLETIM:* {self._format_message_value(data.get('boletim'))}",
            f"*NATUREZA:* {self._format_message_value(data.get('natureza'))}",
            f"*OBSERVACAO:* {self._format_message_value(data.get('observacao'))}",
        ]
        return "\n".join(lines)

    def _open_message_text(self, titulo, conteudo):
        """Gera arquivo de texto temporario e abre no app padrao para copia."""
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        file_name = f"{titulo}_{stamp}.txt"
        file_path = Path(gettempdir()) / file_name
        file_path.write_text(conteudo, encoding="utf-8")

        if sys.platform.startswith("win"):
            os.startfile(str(file_path))  # type: ignore[attr-defined]
            return
        webbrowser.open(file_path.as_uri())

    def criar_talao(self):
        """Processa criacao de novo talao a partir do formulario principal."""
        data = self._collect_form_data()

        try:
            normalized, missing, now = self.talao_service.prepare_new_talao(data)
        except ValueError as exc:
            messagebox.showwarning("Validação", str(exc))
            return

        if missing:
            messagebox.showwarning("Validação", "Campos obrigatórios:\n- " + "\n- ".join(missing))
            return

        intervalo = self.intervalo_map.get(self.alerta_var.get(), DEFAULT_ALERT_INTERVAL_MIN)

        try:
            novo_talao = self.repo.insert_talao(normalized, intervalo)
            messagebox.showinfo("Sucesso", f"Talão {format_talao(now.year, novo_talao)} registrado com status monitorado.")
            self._set_defaults()
            self.refresh_tree()
        except DuplicateTalaoError as exc:
            messagebox.showwarning("Conflito de numeração", str(exc))
            self.refresh_tree()
        except Exception:
            logger.exception("Falha ao gravar novo talão")
            messagebox.showerror("Erro", "Falha ao gravar talão. Verifique os dados e tente novamente.")

    def editar_selecionado(self):
        """Abre a janela de edicao para o talao selecionado na grade."""
        selected = self.tree.selection()
        if not selected:
            messagebox.showinfo("Info", "Selecione um talão na lista.")
            return

        item_id = selected[0]
        valores = self.tree.item(item_id).get("values", [])
        status_atual = str(valores[4]).strip().upper() if len(valores) >= 5 else ""
        if self.alerta_service.is_edit_blocked_status(status_atual):
            messagebox.showinfo("Info", "Talões finalizados ou cancelados não podem ser editados.")
            return

        talao_id = int(item_id)
        intervalo = self.repo.get_monitoring_interval(talao_id)
        if intervalo is None:
            intervalo = DEFAULT_ALERT_INTERVAL_MIN
        TalaoEditor(
            self.root,
            self.repo,
            self.talao_service,
            self.alerta_service,
            talao_id,
            intervalo,
            self.refresh_tree,
        )

    def abrir_relatorios(self):
        """Abre janela modal de relatorios por periodo."""
        RelatorioPeriodoWindow(self.root, self.repo)

    def abrir_backup(self):
        """Abre janela modal de backup anual."""
        BackupAnoWindow(self.root, self.repo)

    def gerar_mensagem_whatsapp_selecionado(self):
        """Gera template de mensagem WhatsApp para o talao selecionado."""
        selected = self.tree.selection()
        if not selected:
            messagebox.showinfo("Info", "Selecione um talão na lista.")
            return

        talao_id = int(selected[0])
        try:
            record = self.repo.get_talao(talao_id)
            if not record:
                messagebox.showwarning("WhatsApp", "Talão não encontrado para gerar mensagem.")
                return
            mensagem = self._build_whatsapp_message(record.get("ano"), record.get("talao"), record)
            self._open_message_text("mensagem_whatsapp_talao", mensagem)
        except Exception:
            logger.exception("Falha ao gerar mensagem WhatsApp do talão %s", talao_id)
            messagebox.showerror("Erro", "Falha ao gerar mensagem para WhatsApp.")

    def refresh_tree(self, silent=False):
        """Recarrega a grade principal com os taloes visiveis."""
        for iid in self.tree.get_children():
            self.tree.delete(iid)

        try:
            rows = self.repo.list_initial_taloes()
        except Exception:
            logger.exception("Falha ao carregar talões")
            if not silent:
                messagebox.showerror("Erro", "Falha ao carregar talões.")
            return

        for row in rows:
            talao_id, ano, talao, boletim, delegacia, natureza, status = row
            self.tree.insert(
                "",
                "end",
                iid=str(talao_id),
                values=(format_talao(ano, talao), boletim or "", delegacia or "", natureza or "", status),
                tags=(status,),
            )

        self._refresh_proximo_talao()

    def _auto_refresh(self):
        """Executa atualizacao periodica silenciosa da grade."""
        self.refresh_tree(silent=True)
        self.root.after(self.AUTO_REFRESH_MS, self._auto_refresh)

    def _has_active_modal(self):
        """Verifica se existe janela modal ativa bloqueando foco."""
        for child in self.root.winfo_children():
            if isinstance(child, tk.Toplevel) and child.winfo_exists():
                if child.grab_current() is child:
                    return True
        return False

    def processar_alertas(self):
        """Processa alertas vencidos de monitoramento em ciclos."""
        if self._has_active_modal():
            self.root.after(self.ALERT_POLL_MS, self.processar_alertas)
            return

        try:
            due_rows = self.repo.list_due_monitoring()
        except Exception:
            logger.exception("Falha ao consultar alertas de monitoramento")
            self.root.after(self.ALERT_POLL_MS, self.processar_alertas)
            return

        # Processa apenas um alerta por ciclo para evitar sequência de pop-ups
        # e garantir foco no preenchimento/validação do talão em questão.
        for row in due_rows:
            talao_id, intervalo_min, ano, talao, boletim, status = row
            if not self.alerta_service.is_monitorado(status):
                continue

            self.root.bell()
            pergunta = self.alerta_service.build_monitoring_question(ano, talao, boletim)
            confirmar = messagebox.askyesno("Alerta de monitoramento", pergunta)

            try:
                if confirmar:
                    self._tentar_finalizar_por_alerta(talao_id, intervalo_min)
                else:
                    self.repo.postpone_monitoring(talao_id, intervalo_min)
            except Exception:
                logger.exception("Falha ao processar alerta do talão %s", talao_id)
                messagebox.showerror("Erro", "Falha ao processar alerta de monitoramento.")
            break

        self.root.after(self.ALERT_POLL_MS, self.processar_alertas)

    def _tentar_finalizar_por_alerta(self, talao_id, intervalo_min):
        """Tenta finalizar talao via alerta, com validacoes e confirmacoes."""
        record = self.repo.get_talao(talao_id)
        if not record:
            return

        try:
            normalized, missing = self.talao_service.prepare_finalize_from_record(record)
        except ValueError as exc:
            messagebox.showwarning("Validação", str(exc))
            self.repo.postpone_monitoring(talao_id, intervalo_min)
            TalaoEditor(
                self.root,
                self.repo,
                self.talao_service,
                self.alerta_service,
                talao_id,
                intervalo_min,
                self.refresh_tree,
            )
            return

        if missing:
            messagebox.showwarning(
                "Pendências",
                "Não foi possível finalizar automaticamente. Campos obrigatórios ausentes:\n- "
                + "\n- ".join(missing),
            )
            self.repo.postpone_monitoring(talao_id, intervalo_min)
            TalaoEditor(
                self.root,
                self.repo,
                self.talao_service,
                self.alerta_service,
                talao_id,
                intervalo_min,
                self.refresh_tree,
            )
            return

        confirmar_envio = messagebox.askyesno(
            "Confirmação obrigatória",
            self.alerta_service.build_final_boletim_confirmation_question(),
            parent=self.root,
        )
        if not confirmar_envio:
            self.repo.postpone_monitoring(talao_id, intervalo_min)
            messagebox.showinfo(
                "Status mantido",
                "O talão permanecerá como MONITORADO porque o boletim finalizado ainda não foi enviado.",
                parent=self.root,
            )
            self.refresh_tree()
            return

        try:
            self.repo.update_talao(
                talao_id,
                normalized,
                intervalo_min,
                expected_updated_at=record.get("atualizado_em"),
            )
            self.refresh_tree()
        except ConcurrencyError as exc:
            messagebox.showwarning("Conflito de edição", str(exc))
            self.refresh_tree()
        except Exception:
            logger.exception("Falha ao finalizar talão %s", talao_id)
            messagebox.showerror("Erro", "Falha ao finalizar talão.")


def build_root():
    """Cria janela raiz usando CTk quando disponivel, senao Tk padrao."""
    if ctk is not None:
        ctk.set_appearance_mode("light")
        ctk.set_default_color_theme("blue")
        return ctk.CTk()
    return tk.Tk()
